import csv,logging
from string import ascii_letters,digits,ascii_uppercase
from shutil import copy2,copy,copyfile
from pathlib import Path
from os import stat as os_stat
from os import remove,listdir,mkdir
from os.path import getsize,getmtime,getctime,isdir,exists
from typing import Any
from secrets import choice
from time import ctime
from decimal import Decimal
# installed modules
# OpenPyXL
openpyxl_imported = True
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    def adjust_column_width(worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            worksheet.column_dimensions[column].width = (max_length - 5) * 1.2
except ImportError:
    openpyxl_imported = False
except ModuleNotFoundError:
    openpyxl_imported = False

unc_path = lambda given_path : str(Path(given_path).resolve()).replace('\\','/')
getCreatedDate = lambda item_path : ctime(getctime(item_path))
getModifiedDate = lambda item_path : ctime(getmtime(item_path))

# This attempts to stop logging messages from installed modules from displaying.
# However, depending on how said modules handle logging, logging messages may
# still end up being displayed for users.
if logging.root.manager.disable != 50:
    root = logging.getLogger()
    for h in tuple(root.handlers):
        root.removeHandler(h)
    root.propagate = False
    try:
        logging.disable(logging.CRITICAL)
    except Exception:
        try:
            logging.disable(logging.FATAL)
        except Exception:
            pass

alnums = f'{ascii_letters}{digits}'

def randstr(string_length : int = 8) -> str:
    '''
    Generate a random string of text of a length specified by string_length
    '''

    return ''.join(choice(alnums) for _ in range(string_length))

def createCopy(source_path : str, dest_path : str) -> None:
    '''
    Generate a copy of a pre-existing file.
    '''

    try:
        copy2(source_path,dest_path)
    except Exception:
        try:
            copy(source_path,dest_path)
        except Exception:
            copyfile(source_path,dest_path)

    return None

def backupGen(original_path : str, backup_path : str) -> None:

    try:
        copy2(original_path,backup_path)
    except Exception:
        try:
            copy(original_path,backup_path)
        except Exception:
            copyfile(original_path,backup_path)

    return None

def getSizeOfItem(item_path : str) -> int | None:
    '''
    Obtain the size of a file in bytes.
    '''

    try: return getsize(item_path)
    except Exception: pass
    try: return os_stat(file_path).st_size
    except Exception: pass
    try: return Path(file_path).stat().st_size
    except Exception: pass

    return None

def getBaselineMetadata(item_path : str) -> tuple[str] | None:
    '''
    Obtain basic metadata from item.
    '''
    baseline_metadata = []
    try:
        baseline_metadata.append(getModifiedDate(item_path)[4:])
        baseline_metadata.append(getCreatedDate(item_path)[4:])
        if (item_size := getSizeOfItem(item_path)) is None:
            return None
        else:
           baseline_metadata.append(str(item_size))
    except Exception:
        return None

    return tuple(baseline_metadata)

def determineEntityType(entity_path : str) -> str | None:

    if not exists(entity_path):
        return None

    if entity_path.lower().endswith('.gdb') and isdir(entity_path):
        return 'GDB'
    elif entity_path.lower().endswith('.shp'):
        return 'SHP'
    elif entity_path.lower().endswith('.txt'):
        return 'TXT'
    elif entity_path.lower().endswith('.pdf'):
        return 'PDF'
    elif entity_path[entity_path.rfind(".")+1:].lower() in ('jpg','jpeg','png','tif','tiff','webp','bmp','dib','icns','ico','jp2','j2k','jpx','pcx','tga','xbm'):
        return 'IMG'
    else:
        return None

def genDuplicateFinderResultFile(found_duplicates : tuple[str], output_type : str, output_location : str, output_name : str, csv_field_size_limit : int, csv_delimiter : str, overwriteOutput : bool, valid_image_types : set) -> None:

    csv.field_size_limit(csv_field_size_limit)

    if output_location in (None,'') or not isinstance(output_location,str):
        user_path = str(Path.home()).replace('\\','/')
        output_location = f'{user_path}/Documents'
    elif not exists((output_location := output_location.strip())):
        if exists(output_location.replace('\\','/')[:output_location.rfind("/")]):
            mkdir(output_location)
        else:
            user_path = str(Path.home()).replace('\\','/')
            output_location = f'{user_path}/Documents'
    elif '\\' in output_location:
        output_location = output_location.replace('\\','/')

    if output_type in ('excel','xlsx'):
        output_suffix = '_cf_duplicate_findings.xlsx'
    elif output_type == 'csv':
        output_suffix = '_cf_duplicate_findings.csv'
    else:
        output_suffix = '_cf_duplicate_findings.txt'

    if output_name in (None,'') or not isinstance(output_name,str):
        output_name = randstr()

    if exists((output_path := f'{output_location}/{output_name}{output_suffix}')):
        if overwriteOutput:
            remove(output_path)
        else:
            neo_output_name = f"{output_name}{output_suffix}"
            existing_files = set(listdir(output_location))
            while neo_output_name in existing_files:
                neo_output_name = f'{output_name}_{randstr()}{output_suffix}'
            output_path = f'{output_location}/{neo_output_name}'
            del existing_files ; del neo_output_name

    del output_suffix

    if output_path.endswith('.xlsx') and openpyxl_imported:
        wb = Workbook()
        organized_files = {}
        counter = {}
        for n in range(len(found_duplicates)):
            if not '.' in found_duplicates[n][0]:
                if 'alia' in organized_files.keys():
                    organized_files['alia'].append(tuple(sorted([item for item in found_duplicates[n]])))
                    counter['alia'] += len(organized_files['alia'][-1])
                else:
                    organized_files['alia'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                    counter['alia'] = len(organized_files['alia'][0])
            else:
                match (extension := found_duplicates[n][0][found_duplicates[n][0].rfind(".")+1:].lower()):
                    case 'txt':
                        if 'txt' in organized_files.keys():
                            organized_files['txt'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            counter['txt'] += len(organized_files['txt'][-1])
                        else:
                            organized_files['txt'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            counter['txt'] = len(organized_files['txt'][0])
                    case 'pdf':
                        if 'pdf' in organized_files.keys():
                            organized_files['pdf'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            counter['pdf'] += len(organized_files['pdf'][-1])
                        else:
                            organized_files['pdf'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            counter['pdf'] = len(organized_files['pdf'][0])
                    case 'shp':
                        if 'shp' in organized_files.keys():
                            organized_files['shp'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            counter['shp'] += len(organized_files['shp'][-1])
                        else:
                            organized_files['shp'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            counter['shp'] = len(organized_files['shp'][0])
                    case 'docx':
                        if 'doc' in organized_files.keys():
                            organized_files['doc'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            counter['doc'] += len(organized_files['doc'][-1])
                        else:
                            organized_files['doc'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            counter['doc'] = len(organized_files['doc'][0])
                    case 'gdb':
                        if 'gdb' in organized_files.keys():
                            organized_files['gdb'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            counter['gdb'] += len(organized_files['gdb'][-1])
                        else:
                            organized_files['gdb'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            counter['gdb'] = len(organized_files['gdb'][0])
                    case _:
                        if extension in valid_image_types:
                            if 'img' in organized_files.keys():
                                organized_files['img'].append(tuple(sorted([item for item in found_duplicates[n]])))
                                counter['img'] += len(organized_files['img'][-1])
                            else:
                                organized_files['img'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                                counter['img'] = len(organized_files['img'][0])
                        elif 'alia' in organized_files.keys():
                            organized_files['alia'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            counter['alia'] += len(organized_files['alia'][-1])
                        else:
                            organized_files['alia'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            counter['alia'] = len(organized_files['alia'][0])
        suffix_association = {'gdb' : 'File Geodatabases', 'img' : 'Images', 'pdf' : 'PDFs', 'shp' : 'ShapeFiles', 'txt' : 'Text Files', 'doc' : 'Word Documents', 'alia' : 'Questionable'}
        upper_letters = tuple(ascii_uppercase)
        for entity_type in tuple(counter.keys()):
            if counter[entity_type] <= 26:
                columns = upper_letters[:counter[entity_type]]
                wb.create_sheet((worksheet_name := suffix_association[entity_type]))
                ws = wb[worksheet_name]
                for n in range(len(organized_files[entity_type])):
                    ws[f'{columns[n]}1'] = f"Match #{n+1}"
                    for x in range(len(organized_files[entity_type][n])):
                        ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                adjust_column_width(ws)
            elif counter[entity_type] <= 702:
                columns = tuple(list(upper_letters) + [f'{a}{b}' for a in upper_letters for b in upper_letters])[:counter[entity_type]]
                wb.create_sheet((worksheet_name := suffix_association[entity_type]))
                ws = wb[worksheet_name]
                for n in range(len(organized_files[entity_type])):
                    ws[f'{columns[n]}1'] = f"Match #{n+1}"
                    for x in range(len(organized_files[entity_type][n])):
                        ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                adjust_column_width(ws)
            elif counter[entity_type] <= 16_384:
                columns = tuple(list(upper_letters) + [f'{a}{b}' for a in upper_letters for b in upper_letters] + [f'{a}{b}{c}' for a in upper_letters for b in upper_letters for c in upper_letters])[:counter[entity_type]]
                wb.create_sheet((worksheet_name := suffix_association[entity_type]))
                ws = wb[worksheet_name]
                for n in range(len(organized_files[entity_type])):
                    ws[f'{columns[n]}1'] = f"Match #{n+1}"
                    for x in range(len(organized_files[entity_type][n])):
                        ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                adjust_column_width(ws)
            else:
                columns = tuple(list(upper_letters) + [f'{a}{b}' for a in upper_letters for b in upper_letters] + [f'{a}{b}{c}' for a in upper_letters for b in upper_letters for c in upper_letters])[:16_384]
                num = counter[entity_type]
                iteration_num = 1
                match_counter = 1
                while num > 16_384:
                    wb.create_sheet((sheet_title := f'{suffix_association[entity_type]} {iteration_num}'))
                    ws = wb[sheet_title]
                    for n in range(16_384):
                        ws[f'{columns[n]}1'] = f"Match #{match_counter}"
                        match_counter += 1
                        for x in range(16_384):
                            ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                    adjust_column_width(ws)
                    organized_files[entity_type] = tuple(list(organized_files[entity_type])[16_384:])
                    iteration_num += 1
                    num -= 16_384
                if len(organized_files[entity_type]) <= 26:
                    columns = upper_letters[:len(organized_files[entity_type])]
                    wb.create_sheet(sheet_title = f'{suffix_association[entity_type]} {iteration_num}')
                    ws = wb[sheet_title]
                    for n in range(len(columns)):
                        ws[f'{columns[n]}1'] = f"Match # {match_counter}"
                        match_counter += 1
                        for x in range(len(organized_files[entity_type][n])):
                            ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                    adjust_column_width(ws)
                elif len(organized_files[entity_type]) <= 702:
                    columns = tuple(list(upper_letters) + [f'{a}{b}' for a in upper_letters for b in upper_letters])[:len(organized_files[entity_type])]
                    wb.create_sheet(sheet_title = f'{suffix_association[entity_type]} {iteration_num}')
                    ws = wb[sheet_title]
                    for n in range(len(columns)):
                        ws[f'{columns[n]}1'] = f"Match # {match_counter}"
                        match_counter += 1
                        for x in range(len(organized_files[entity_type][n])):
                            ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                    adjust_column_width(ws)
                else:
                    columns = tuple(list(upper_letters) + [f'{a}{b}' for a in upper_letters for b in upper_letters] + [f'{a}{b}{c}' for a in upper_letters for b in upper_letters for c in upper_letters])[:len(organized_files[entity_type])]
                    wb.create_sheet(sheet_title = f'{suffix_association[entity_type]} {iteration_num}')
                    ws = wb[sheet_title]
                    for n in range(len(columns)):
                        ws[f'{columns[n]}1'] = f"Match # {match_counter}"
                        match_counter += 1
                        for x in range(len(organized_files[entity_type][n])):
                            ws[f'{columns[n]}{x+2}'] = organized_files[entity_type][n][x]
                    adjust_column_width(ws)
        try:
            del wb['Sheet']
        except Exception:
            try:
                del wb['sheet']
            except Exception:
                pass
        wb.save(output_path)
        wb.close()
    else:
        organized_files = {}
        if output_path.endswith('.csv'):
            for n in range(len(found_duplicates)):
                if not '.' in found_duplicates[n][0]:
                    if not 'alia' in organized_files.keys():
                        organized_files['alia'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                    else:
                        organized_files['alia'].append(tuple(sorted([item for item in found_duplicates[n]])))
                else:
                    match (extension := found_duplicates[n][0][found_duplicates[n][0].rfind(".")+1:].lower()):
                        case 'txt':
                            if not 'txt' in organized_files.keys():
                                organized_files['txt'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['txt'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'pdf':
                            if not 'pdf' in organized_files.keys():
                                organized_files['pdf'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['pdf'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'shp':
                            if not 'shp' in organized_files.keys():
                                organized_files['shp'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['shp'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'gdb':
                            if not 'gdb' in organized_files.keys():
                                organized_files['gdb'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['gdb'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'docx':
                            if not 'doc' in organized_files.keys():
                                organized_files['doc'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['doc'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case _:
                            if extension in valid_image_types:
                                if not 'img' in organized_files.keys():
                                    organized_files['img'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                                else:
                                    organized_files['img'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            elif not 'alia' in organized_files.keys():
                                organized_files['alia'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['alia'].append(tuple(sorted([item for item in found_duplicates[n]])))
            with open(output_path,'w') as cf:
                csv_writer = csv.writer(cf,delimiter=csv_delimiter)
                csv_writer.writerow(["Explicit Paths to Potenial Matching Duplicates"])
                for entity_key in organized_files.keys():
                    for grouping in organized_files[entity_key]:
                        csv_writer.writerow(grouping)
        else:
            for n in range(len(found_duplicates)):
                if not '.' in found_duplicates[n][0]:
                    if not 'alia' in organized_files.keys():
                        organized_files['alia'] = tuple(sorted([item for item in found_duplicates[n]]))
                    else:
                        organized_files['alia'].append(tuple(sorted([item for item in found_duplicates[n]])))
                else:
                    match (extension := found_duplicates[n][0][found_duplicates[n][0].rfind(".")+1:].lower()):
                        case 'txt':
                            if not 'txt' in organized_files.keys():
                                organized_files['txt'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['txt'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'pdf':
                            if not 'pdf' in organized_files.keys():
                                organized_files['pdf'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['pdf'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'shp':
                            if not 'shp' in organized_files.keys():
                                organized_files['shp'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['shp'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'gdb':
                            if not 'gdb' in organized_files.keys():
                                organized_files['gdb'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['gdb'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case 'docx':
                            if not 'doc' in organized_files.keys():
                                organized_files['doc'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['doc'].append(tuple(sorted([item for item in found_duplicates[n]])))
                        case _:
                            if extension in valid_image_types:
                                if not 'img' in organized_files.keys():
                                    organized_files['img'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                                else:
                                    organized_files['img'].append(tuple(sorted([item for item in found_duplicates[n]])))
                            elif not 'alia' in organized_files.keys():
                                organized_files['alia'] = [tuple(sorted([item for item in found_duplicates[n]]))]
                            else:
                                organized_files['alia'].append(tuple(sorted([item for item in found_duplicates[n]])))
            output_path = f'{output_path[:output_path.rfind(".")]}.txt'
            with open(output_path,'w',encoding='utf-8') as tf:
                tf.write("Explicit Paths to Potenial Matching Duplicates")
                for entity_key in organized_files.keys():
                    for grouping in organized_files[entity_key]:
                        tf.write("\n%s" % ("|".join(grouping)))

    return None

def genSearchQueryResultFile(found_matches : tuple[str], output_type : str, output_location : str, output_name : str, csv_field_size_limit : int, csv_delimiter : str, overwriteOutput : bool, valid_image_types : set) -> None:

    csv.field_size_limit(csv_field_size_limit)

    if output_location in (None,'') or not isinstance(output_location,str):
        user_path = str(Path.home()).replace('\\','/')
        output_location = f'{user_path}/Documents'
    elif not exists((output_location := output_location.strip())):
        if exists(output_location.replace('\\','/')[:output_location.rfind("/")]):
            mkdir(output_location)
        else:
            user_path = str(Path.home()).replace('\\','/')
            output_location = f'{user_path}/Documents'
    elif '\\' in output_location:
        output_location = output_location.replace('\\','/')

    if output_type in ('excel','xlsx'):
        output_suffix = '.xlsx'
    elif output_type == 'csv':
        output_suffix = '.csv'
    else:
        output_suffix = '.txt'

    if output_name in (None,'') or not isinstance(output_name,str):
        output_name = randstr()

    if exists((output_path := f'{output_location}/{output_name}{output_suffix}')):
        if overwriteOutput:
            remove(output_path)
        else:
            neo_output_name = f"{output_name}{output_suffix}"
            existing_files = set(listdir(output_location))
            while neo_output_name in existing_files:
                neo_output_name = f'{output_name}_{randstr()}{output_suffix}'
            output_path = f'{output_location}/{neo_output_name}'
            del existing_files ; del neo_output_name

    del output_suffix

    organized_files = {}
    for item in found_matches:
        match (extension := item[item.rfind('.')+1:].lower()):
            case 'txt':
                if not 'txt' in organized_files.keys():
                    organized_files['txt'] = [item]
                else:
                    organized_files['txt'].append(item)
            case 'pdf':
                if not 'pdf' in organized_files.keys():
                    organized_files['pdf'] = [item]
                else:
                    organized_files['pdf'].append(item)
            case 'shp':
                if not 'shp' in organized_files.keys():
                    organized_files['shp'] = [item]
                else:
                    organized_files['shp'].append(item)
            case 'gdb':
                if not 'gdb' in organized_files.keys():
                    organized_files['gdb'] = [item]
                else:
                    organized_files['gdb'].append(item)
            case 'docx':
                if not 'doc' in organized_files.keys():
                    organized_files['doc'] = [item]
                else:
                    organized_files['doc'].append(item)
            case _:
                if extension in valid_image_types:
                    if not 'img' in organized_files.keys():
                        organized_files['img'] = [item]
                    else:
                        organized_files['img'].append(item)
                else:
                    if not 'alia' in organized_files.keys():
                        organized_files['alia'] = [item]
                    else:
                        organized_files['alia'].append(item)

    for entity_key in tuple(organized_files.keys()):
        organized_files[entity_key] = tuple(sorted(organized_files[entity_key]))
    if output_path.endswith('.xlsx') and openpyxl_imported:
        key_association = {"File Geodatabases":"gdb","Images":"img","PDFs":"pdf","ShapeFiles":"shp","Text Files":"txt","Word Documents":"doc","Miscellaneous":"alia"}
        wb = Workbook()
        for worksheet_name in ("File Geodatabases","Images","PDFs","ShapeFiles","Text Files","Word Documents","Miscellaneous"):
            if key_association[worksheet_name] in organized_files.keys():
                wb.create_sheet(worksheet_name)
                ws = wb[worksheet_name]
                for n in range(len(organized_files[key_association[worksheet_name]])):
                    ws[f"A{n+1}"] = organized_files[key_association[worksheet_name]][n]
                adjust_column_width(ws)
        try:
            del wb['Sheet']
        except Exception:
            try:
                del wb['sheet']
            except Exception:
                pass
        wb.save(output_path)
        wb.close()
    elif output_path.endswith('.csv'):
        with open(output_path,'w') as cf:
            csv_writer = csv.writer(cf,delimiter=csv_delimiter)
            csv_writer.writerow(["Explicit Path to Valid Item"])
            for entity_key in organized_files.keys():
                for entity_path in organized_files[entity_key]:
                    csv_writer.writerow([entity_path])
    else:
        output_path = f'{output_path[:output_path.rfind(".")]}.txt'
        with open(output_path,'w',encoding='utf-8') as tf:
            tf.write("Explicit Path to Valid Item")
            for entity_key in organized_files.keys():
                for entity_path in organized_files[entity_key]:
                    tf.write(f'\n{entity_path}')

    return None

def forbidden_dirs() -> set[str]:
    '''
    This is a baseline protection against malicious executions of Chloe Felina on Windows OS.
    '''

    default_things = ("Downloads","Documents","AppData","Contacts","Favorites","Links","Music","Pictures","Saved Games","Searches","Videos")

    bad_items = {"C:",}
    for item in tuple(listdir('C:')):
        if isdir(f'C:/Users/{item}'):
            bad_items.add(f'C:/{item}')
    try:
        for registered_user in tuple([item for item in tuple(listdir("C:/Users")) if isdir(f"C:/Users/{item}")]):
            bad_items.add(f'C:/Users/{registered_user}')
            for default_thing in default_things:
                if exists(f'C:/Users/{registered_user}/{default_thing}'):
                    bad_items.add(f'C:/Users/{registered_user}/{default_thing}')
                    for item in tuple(listdir(f'C:/Users/{registered_user}/{default_thing}')):
                        if isdir(f'C:/Users/{registered_user}/{default_thing}/{item}'):
                            bad_items.add(f'C:/Users/{registered_user}/{default_thing}/{item}')
    except Exception:
        # In case there is somehow not a C:/Users directory.
        pass
    for main_dir in ("Program Files","Program Files (x86)","Recovery","System.sav","Temp"):
        try:
            for item in tuple(listdir(f'C:/{main_dir}')):
                if isdir(f'C:/{main_dir}/{item}'):
                    bad_items.add(f'C:/{main_dir}/{item}')
        except Exception:
            pass

    return bad_items
