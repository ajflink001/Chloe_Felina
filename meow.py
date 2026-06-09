import csv
from string import ascii_letters,digits
from shutil import copy2,copy,copyfile
from pathlib import Path
from os import stat as os_stat
from os import getlogin,remove,listdir
from os.path import getsize,getmtime,getctime,isdir,exists
from typing import Union,Any
from secrets import choice
from time import ctime
from decimal import Decimal
# installed modules
# PyWin32
pywin32_imported = True
try: from win32api import GetLogicalDriveStrings
except ImportError: pywin32_imported = False
except ModuleNotFoundError: pywin32_imported = False
# OpenPyXL
openpyxl_imported = True
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    def adjust_column_width(worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            worksheet.column_dimensions[column].width = (max_length - 5) * 1.2
except ImportError:
    openpyxl_imported = False
except ModuleNotFoundError:
    openpyxl_imported = False

unc_path = lambda given_path : str(Path(given_path).resolve()).replace('\\','/')
getCreatedDate = lambda item_path : ctime(getctime(item_path))
getModifiedDate = lambda item_path : ctime(getmtime(item_path))

alnums = f'{ascii_letters}{digits}'

if pywin32_imported:
    comp_drives = {unc_path(drive) : drive.replace('\\','/') for drive in GetLogicalDriveStrings().split('\000')[:-1] if drive != 'C:\\'}
    explicit_drives = tuple(comp_drives.keys())
    associated_drives = {comp_drives[item] : item for item in explicit_drives}

    del comp_drives

def randstr(string_length : int = 8) -> str:
    '''
    Generate a random string of text of a length specified by string_length
    '''

    return ''.join(choice(alnums) for _ in range(string_length))

def createCopy(source_path : str, dest_path : str) -> None:
    '''
    Generate a copy of a pre-existing file.
    '''

    try:
        copy2(source_path,dest_path)
    except Exception:
        try:
            copy(source_path,dest_path)
        except Exception:
            copyfile(source_path,dest_path)

    return None

def backupGen(original_path : str, backup_path : str) -> None:

    try:
        copy2(original_path,backup_path)
    except Exception:
        try:
            copy(original_path,backup_path)
        except Exception:
            copyfile(original_path,backup_path)

    return None

def getSizeOfItem(item_path : str) -> Union[int,None]:
    '''
    Obtain the size of a file in bytes.
    '''

    try: return getsize(item_path)
    except Exception: pass
    try: return os_stat(file_path).st_size
    except Exception: pass
    try: return Path(file_path).stat().st_size
    except Exception: pass

    return None

def getBaselineMetadata(item_path : str) -> Union[tuple,None]:
    '''
    Obtain basic metadata from item.
    '''
    baseline_metadata = []
    try:
        baseline_metadata.append(getModifiedDate(item_path)[4:])
        baseline_metadata.append(getCreatedDate(item_path)[4:])
        if (item_size := getSizeOfItem(item_path)) is None:
            return None
        else:
           baseline_metadata.append(str(item_size))
    except Exception:
        return None

    return tuple(baseline_metadata)

def determineEntityType(entity_path : str) -> Union[str,None]:

    if not exists(entity_path):
        return None

    if entity_path.lower().endswith('.gdb') and isdir(entity_path):
        return 'GDB'
    elif entity_path.lower().endswith('.shp'):
        return 'SHP'
    elif entity_path.lower().endswith('.txt'):
        return 'TXT'
    elif entity_path.lower().endswith('.pdf'):
        return 'PDF'
    elif entity_path[entity_path.rfind(".")+1:].lower() in ('jpg','jpeg','png','tif','tiff','webp'):
        return 'IMG'
    else:
        return None

def convertValueToBytes(value : Union[int,float], unit : str) -> int:
    '''
    This will result in inaccurate values if the resulting value exceeds or
    approaches:
    18_446_744_073_709_551_615 (i.e., unsigned 2^64 minus 1)
    '''

    # Zettabytes to quettabytes are excluded due to how extremely niche
    # such high values would be in practical use-cases. They are left in
    # case some methodology is determined on how to determine such extremely
    # high values with true accuracy.

    match unit.lower():
        case 'kb' | 'kilobytes' | 'kilobyte':
            return round(float(Decimal(value) * Decimal(1_000)))
        case 'kib' | 'kilibytes' | 'kilibyte':
            return round(float(Decimal(value) * Decimal(1_024)))
        case 'mb' | 'megabytes' | 'megabyte':
            return round(float(Decimal(value) * Decimal(1_000_000)))
        case 'mib' | 'mebibytes' | 'mebibyte':
            return round(float(Decimal(value) * Decimal(1_048_576)))
        case 'gb' | 'gigabytes' | 'gigabyte':
            return round(float(Decimal(value) * Decimal(1_000_000_000)))
        case 'gib' | 'gibibytes' | 'gibibyte':
            return round(float(Decimal(value) * Decimal(1_073_741_824)))
        case 'tb' | 'terabytes' | 'terabyte':
            return round(float(Decimal(value) * Decimal(1_000_000_000_000)))
        case 'tib' | 'tebibytes' | 'tebibyte':
            return round(float(Decimal(value) * Decimal(1_099_511_627_776)))
        case 'pb' | 'petabytes' | 'petabytes':
            return round(float(Decimal(value) * Decimal(1_000_000_000_000_000)))
        case 'pib' | 'pebibytes' | 'pebibyte':
            return round(float(Decimal(value) * Decimal(1_125_899_906_842_624)))
        case 'eb' | 'exabytes' | 'exabyte':
            return round(float(Decimal(value) * Decimal(1_000_000_000_000_000_000)))
        case 'eib' | 'exbibytes' | 'exbibyte':
            return round(float(Decimal(value) * Decimal(1_152_921_504_606_845_976)))
        case 'zb' | 'zettabytes' | 'zettabyte':
            return round(value)
        case 'zib' | 'zebibytes' | 'zebibyte':
            return round(value)
        case 'yb' | 'yottabytes' | 'yottabyte':
            return round(value)
        case 'yib' | 'yobibytes' | 'yobibyte':
            return round(value)
        case 'rb' | 'ronnabytes' | 'ronnabyte':
            return round(value)
        case 'rib' | 'robibytes' | 'robibyte':
            return round(value)
        case 'qb' | 'quettabytes' | 'quettabyte':
            return round(value)
        case 'qib' | 'quebibytes' | 'quebibyte':
            return round(value)
        case _:
            return round(value)

def genSearchQueryResultFile(found_matches : tuple, output_type : str, output_location : str, output_name : str, csv_field_size_limit : int, csv_delimiter : str, csv_quotechar : str, csv_quoting_minimal : int, csv_newline : str, overwriteOutput : bool) -> None:

    csv.field_size_limit(csv_field_size_limit)

    if output_location in (None,''):
        output_location = f'C:/Users/{getlogin()}/Documents'
    elif not exists((output_location := output_location.strip())):
        output_location = f'C:/Users/{getlogin()}/Documents'
    elif '\\' in output_location:
        output_location = output_location.replace('\\','/')

    if output_type in ('excel','xlsx'):
        output_suffix = '.xlsx'
    elif output_type == 'csv':
        output_suffix = '.csv'
    else:
        output_suffix = '.txt'

    if exists((output_path := f'{output_location}/{output_name}{output_suffix}')):
        if overwriteOutput:
            remove(output_path)
        else:
            neo_output_name = f"{output_name}{output_suffix}"
            existing_files = set(listdir(output_location))
            while neo_output_name in existing_files:
                neo_output_name = f'{output_name}_{randstr()}{output_suffix}'
            output_path = f'{output_location}/{neo_output_name}'
            del existing_files ; del neo_output_name

    del output_suffix

    organized_files = {'gdb':[],'img':[],'pdf':[],'shp':[],'txt':[]}
    for item in found_matches:
        match item[item.rfind('.')+1:].lower():
            case 'txt':
                organized_files['txt'].append(item)
            case 'pdf':
                organized_files['pdf'].append(item)
            case 'shp':
                organized_files['shp'].append(item)
            case 'gdb':
                organized_files['gdb'].append(item)
            case _:
                organized_files['img'].append(item)

    for entity_key in tuple(organized_files.keys()):
        organized_files[entity_key] = tuple(sorted(organized_files[entity_key]))
    if output_path.endswith('.xlsx') and openpyxl_imported:
        key_association = {"File Geodatabases":"gdb","Images":"img","PDFs":"pdf","ShapeFiles":"shp","Text Files":"txt"}
        wb = Workbook()
        for worksheet_name in ("File Geodatabases","Images","PDFs","ShapeFiles","Text Files"):
            wb.create_sheet(worksheet_name)
            ws = wb[worksheet_name]
            for n in range(len(organized_files[key_association[worksheet_name]])):
                ws[f"A{n+1}"] = organized_files[key_association[worksheet_name]][n]
            adjust_column_width(ws)
        try:
            del wb['Sheet']
        except Exception:
            try:
                del wb['sheet']
            except Exception:
                pass
        wb.save(output_path)
        wb.close()
    elif output_path.endswith('.csv'):
        with open(output_path,'w',newline=csv_newline) as cf:
            csv_writer = csv.writer(cf,delimiter=csv_delimiter,quotechar=csv_quotechar,quoting=csv_quoting_minimal)
            csv_writer.writerow(["Explicit Path to Valid Item"])
            for entity_key in organized_files.keys():
                for entity_path in organized_files[entity_key]:
                    csv_writer.writerow([entity_path])
    else:
        if not openpyxl_imported:
            output_path = f'{output_path[:output_path.rfind(".")]}.txt'
        with open(output_path,'w',encoding='utf-8') as tf:
            tf.write("Explicit Path to Valid Item")
            for entity_key in organized_files.keys():
                for entity_path in organized_files[entity_key]:
                    tf.write(f'\n{entity_path}')

    return None

def forbidden_dirs() -> set:
    '''
    This is a baseline protection against malicious executions of Chloe Felina on Windows OS.
    '''

    default_things = ("Downloads","Documents","AppData","Contacts","Favorites","Links","Music","Pictures","Saved Games","Searches","Videos","AppData/Local","AppData/LocalLow","AppData/Roaming")

    bad_items = {"C:",}
    bad_items += {item for item in tuple(listdir("C:")) if isidir(f"C:/{item}")}
    bad_items += (registered_users := {item for item in tuple(listdir("C:/Users")) if isdir(f"C:/Users/{item}")})
    for registered_user in tuple(registered_users):
        bad_items += {item for item in tuple(listdir(f"C:/Users/{registered_user}")) if isdir(f"C:/Users/{registered_user}/{item}")}
        for default_thing in default_things:
            bad_items += {item for item in tuple(listdir(f"C:/Users/{registered_user}/{default_thing}")) if isdir(f"C:/Users/{registered_user}/{default_thing}/{item}")}
    del registered_users
    bad_items += {item for item in tuple(listdir("C:/Program Files")) if isdir(f"C:/Program Files/{item}")}
    bad_items += {item for item in tuple(listdir("C:/Program Files (x86)")) if isdir(f"C:/Program Files (x86)/{item}")}
    bad_items += {item for item in tuple(listdir("C:/ProgramData")) if isdir(f"C:/ProgramData/{item}")}
    bad_items += {item for item in tuple(listdir("C:/Recovery")) if isdir(f"C:/Recovery/{item}")}
    bad_items += {item for item in tuple(listdir("C:/System.Sav")) if isdir(f"C:/System.Sav/{item}")}
    bad_items += {item for item in tuple(listdir("C:/temp_folder")) if isdir(f"C:/Temp/{item}")}

    return bad_items
